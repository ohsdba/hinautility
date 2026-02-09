from flask import Flask, render_template, request, jsonify, send_file
import psycopg2
from psycopg2 import OperationalError, ProgrammingError
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta
import os
import tempfile
import logging
import time
import json
import uuid
import re
from contextlib import contextmanager
import csv
from io import StringIO, BytesIO
import base64
import threading
import time
from functools import wraps
from collections import defaultdict
import hashlib

# ===================== 初始化配置 =====================

# 数据库配置文件锁
DB_CONFIG_LOCK = threading.Lock()
app = Flask(__name__, template_folder='html', static_folder='static')

# 获取路径配置
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 项目根目录为 app.py 所在目录
PROJECT_ROOT = BASE_DIR

# 临时目录
TEMP_DIR = tempfile.gettempdir()
# 数据库配置（优先读取本地配置文件，无则用默认）
DB_CONFIG_FILE = os.path.join(PROJECT_ROOT, "conf", "db_config.json")
DEFAULT_DB_CONFIG = {
    "host": "localhost",
    "port": "5432",
    "user": "postgres",
    "password": "123456",
    "database": "postgres"
}

def get_db_timeout_config():
    """从应用配置中获取数据库超时配置"""
    try:
        # 从APP_CONFIG获取数据库超时配置，如果没有则使用默认值
        # 检查APP_CONFIG是否存在，防止初始化问题
        if 'APP_CONFIG' in globals() and APP_CONFIG:
            statement_timeout = APP_CONFIG.get("db_statement_timeout", 30)
            connect_timeout = APP_CONFIG.get("db_connect_timeout", 10)
        else:
            # 如果APP_CONFIG尚未初始化，则尝试从配置文件加载
            import json
            import os
            app_config_file = os.path.join(PROJECT_ROOT, "conf", "app_config.json")
            if os.path.exists(app_config_file):
                with open(app_config_file, 'r', encoding='utf-8') as f:
                    app_config_loaded = json.load(f)
                statement_timeout = app_config_loaded.get("db_statement_timeout", 30)
                connect_timeout = app_config_loaded.get("db_connect_timeout", 10)
            else:
                # 使用默认值
                statement_timeout = 30
                connect_timeout = 10
        
        return {
            "statement_timeout": statement_timeout,
            "connect_timeout": connect_timeout
        }
    except Exception as e:
        logging.warning(f"获取数据库超时配置失败，使用默认配置: {str(e)}")
        return {
            "statement_timeout": 30,
            "connect_timeout": 10
        }

# 数据库超时配置将在应用配置加载后初始化
DB_TIMEOUT_CONFIG = {}

# 应用配置（优先读取本地配置文件，无则用默认）
APP_CONFIG_FILE = os.path.join(PROJECT_ROOT, "conf", "app_config.json")
# 定义默认应用配置
DEFAULT_APP_CONFIG = dict(
    app_auto_lock_timeout_minutes=30,
    app_auto_lock_reminder_minutes=25,
    app_title="朝阳数据",
    db_statement_timeout=30,
    db_connect_timeout=10,
    app_password="",
    app_max_connections=10,
    app_min_connections=1,
    app_connection_pool_timeout=30,
    app_result_cache_time=3600,
    app_max_result_size=10000,
    app_login_failures_limit=5,
    app_account_lockout_minutes=30,
    app_password_strength_required=False,  # 使用Python布尔值
    app_log_level="INFO",
    app_log_retention_days=30,
    app_audit_logging_enabled=True,  # 使用Python布尔值
    app_theme_color="default",
    app_language="zh-CN",
    app_page_size=50,
    app_auto_save_interval=300,
    app_concurrent_queries=5,
    app_query_queue_size=10,
    app_memory_limit_mb=512,
    app_batch_insert_size=1000,
    app_transaction_timeout=120,
    app_connection_retry_count=3
)

def load_app_config():
    """加载应用配置"""
    try:
        if os.path.exists(APP_CONFIG_FILE):
            with open(APP_CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
                # 确保配置包含必需字段
                app_config = DEFAULT_APP_CONFIG.copy()
                app_config.update(config)
                
                # 验证配置值的合理性
                if 'app_auto_lock_timeout_minutes' in app_config:
                    timeout = app_config['app_auto_lock_timeout_minutes']
                    # 确保超时时间在合理范围内（最小1分钟，最大24小时）
                    if not isinstance(timeout, (int, float)) or timeout < 1 or timeout > 1440:
                        logging.warning(f"自动锁定超时时间超出合理范围，使用默认值: {timeout}")
                        app_config['app_auto_lock_timeout_minutes'] = DEFAULT_APP_CONFIG['app_auto_lock_timeout_minutes']
                
                if 'app_auto_lock_reminder_minutes' in app_config:
                    reminder = app_config['app_auto_lock_reminder_minutes']
                    # 确保提醒时间小于超时时间且在合理范围内
                    timeout = app_config['app_auto_lock_timeout_minutes']
                    if not isinstance(reminder, (int, float)) or reminder <= 0 or reminder >= timeout:
                        logging.warning(f"自动锁定提醒时间不合理，使用默认值: {reminder}")
                        app_config['app_auto_lock_reminder_minutes'] = DEFAULT_APP_CONFIG['app_auto_lock_reminder_minutes']
                
                # 验证数据库超时配置
                if 'db_statement_timeout' in app_config:
                    stmt_timeout = app_config['db_statement_timeout']
                    if not isinstance(stmt_timeout, (int, float)) or stmt_timeout <= 0:
                        logging.warning(f"语句超时时间不合理，使用默认值: {stmt_timeout}")
                        app_config['db_statement_timeout'] = DEFAULT_APP_CONFIG['db_statement_timeout']
                
                if 'db_connect_timeout' in app_config:
                    conn_timeout = app_config['db_connect_timeout']
                    if not isinstance(conn_timeout, (int, float)) or conn_timeout <= 0:
                        logging.warning(f"连接超时时间不合理，使用默认值: {conn_timeout}")
                        app_config['db_connect_timeout'] = DEFAULT_APP_CONFIG['db_connect_timeout']
                
                # 处理应用密码字段
                if 'app_password' not in app_config:
                    app_config['app_password'] = DEFAULT_APP_CONFIG['app_password']
                
                # 验证和设置其他应用配置项
                # 连接池配置
                if 'app_max_connections' in app_config:
                    max_conn = app_config['app_max_connections']
                    if not isinstance(max_conn, (int, float)) or max_conn <= 0:
                        logging.warning(f"最大连接数不合理，使用默认值: {max_conn}")
                        app_config['app_max_connections'] = DEFAULT_APP_CONFIG['app_max_connections']
                
                if 'app_min_connections' in app_config:
                    min_conn = app_config['app_min_connections']
                    if not isinstance(min_conn, (int, float)) or min_conn < 0:
                        logging.warning(f"最小连接数不合理，使用默认值: {min_conn}")
                        app_config['app_min_connections'] = DEFAULT_APP_CONFIG['app_min_connections']
                
                if 'app_connection_pool_timeout' in app_config:
                    pool_timeout = app_config['app_connection_pool_timeout']
                    if not isinstance(pool_timeout, (int, float)) or pool_timeout <= 0:
                        logging.warning(f"连接池超时时间不合理，使用默认值: {pool_timeout}")
                        app_config['app_connection_pool_timeout'] = DEFAULT_APP_CONFIG['app_connection_pool_timeout']
                
                # 结果缓存配置
                if 'app_result_cache_time' in app_config:
                    cache_time = app_config['app_result_cache_time']
                    if not isinstance(cache_time, (int, float)) or cache_time <= 0:
                        logging.warning(f"结果缓存时间不合理，使用默认值: {cache_time}")
                        app_config['app_result_cache_time'] = DEFAULT_APP_CONFIG['app_result_cache_time']
                
                if 'app_max_result_size' in app_config:
                    max_size = app_config['app_max_result_size']
                    if not isinstance(max_size, (int, float)) or max_size <= 0:
                        logging.warning(f"最大结果集大小不合理，使用默认值: {max_size}")
                        app_config['app_max_result_size'] = DEFAULT_APP_CONFIG['app_max_result_size']
                
                # 安全配置
                if 'app_login_failures_limit' in app_config:
                    fail_limit = app_config['app_login_failures_limit']
                    if not isinstance(fail_limit, (int, float)) or fail_limit <= 0:
                        logging.warning(f"登录失败次数限制不合理，使用默认值: {fail_limit}")
                        app_config['app_login_failures_limit'] = DEFAULT_APP_CONFIG['app_login_failures_limit']
                
                if 'app_account_lockout_minutes' in app_config:
                    lockout_minutes = app_config['app_account_lockout_minutes']
                    if not isinstance(lockout_minutes, (int, float)) or lockout_minutes <= 0:
                        logging.warning(f"账户锁定时间不合理，使用默认值: {lockout_minutes}")
                        app_config['app_account_lockout_minutes'] = DEFAULT_APP_CONFIG['app_account_lockout_minutes']
                
                # 日志配置
                if 'app_log_retention_days' in app_config:
                    retention_days = app_config['app_log_retention_days']
                    if not isinstance(retention_days, (int, float)) or retention_days <= 0:
                        logging.warning(f"日志保留天数不合理，使用默认值: {retention_days}")
                        app_config['app_log_retention_days'] = DEFAULT_APP_CONFIG['app_log_retention_days']
                
                # 界面配置
                if 'app_page_size' in app_config:
                    page_size = app_config['app_page_size']
                    if not isinstance(page_size, (int, float)) or page_size <= 0:
                        logging.warning(f"页面大小不合理，使用默认值: {page_size}")
                        app_config['app_page_size'] = DEFAULT_APP_CONFIG['app_page_size']
                
                if 'app_auto_save_interval' in app_config:
                    save_interval = app_config['app_auto_save_interval']
                    if not isinstance(save_interval, (int, float)) or save_interval < 0:
                        logging.warning(f"自动保存间隔不合理，使用默认值: {save_interval}")
                        app_config['app_auto_save_interval'] = DEFAULT_APP_CONFIG['app_auto_save_interval']
                
                # 性能配置
                if 'app_concurrent_queries' in app_config:
                    concurrent_queries = app_config['app_concurrent_queries']
                    if not isinstance(concurrent_queries, (int, float)) or concurrent_queries <= 0:
                        logging.warning(f"并发查询数不合理，使用默认值: {concurrent_queries}")
                        app_config['app_concurrent_queries'] = DEFAULT_APP_CONFIG['app_concurrent_queries']
                
                if 'app_query_queue_size' in app_config:
                    queue_size = app_config['app_query_queue_size']
                    if not isinstance(queue_size, (int, float)) or queue_size <= 0:
                        logging.warning(f"查询队列大小不合理，使用默认值: {queue_size}")
                        app_config['app_query_queue_size'] = DEFAULT_APP_CONFIG['app_query_queue_size']
                
                if 'app_memory_limit_mb' in app_config:
                    memory_limit = app_config['app_memory_limit_mb']
                    if not isinstance(memory_limit, (int, float)) or memory_limit <= 0:
                        logging.warning(f"内存限制不合理，使用默认值: {memory_limit}")
                        app_config['app_memory_limit_mb'] = DEFAULT_APP_CONFIG['app_memory_limit_mb']
                
                if 'app_batch_insert_size' in app_config:
                    batch_size = app_config['app_batch_insert_size']
                    if not isinstance(batch_size, (int, float)) or batch_size <= 0:
                        logging.warning(f"批量插入大小不合理，使用默认值: {batch_size}")
                        app_config['app_batch_insert_size'] = DEFAULT_APP_CONFIG['app_batch_insert_size']
                
                if 'app_transaction_timeout' in app_config:
                    trans_timeout = app_config['app_transaction_timeout']
                    if not isinstance(trans_timeout, (int, float)) or trans_timeout <= 0:
                        logging.warning(f"事务超时时间不合理，使用默认值: {trans_timeout}")
                        app_config['app_transaction_timeout'] = DEFAULT_APP_CONFIG['app_transaction_timeout']
                
                if 'app_connection_retry_count' in app_config:
                    retry_count = app_config['app_connection_retry_count']
                    if not isinstance(retry_count, (int, float)) or retry_count < 0:
                        logging.warning(f"连接重试次数不合理，使用默认值: {retry_count}")
                        app_config['app_connection_retry_count'] = DEFAULT_APP_CONFIG['app_connection_retry_count']
                
                return app_config
        else:
            return DEFAULT_APP_CONFIG
    except Exception as e:
        logging.warning(f"读取应用配置失败，使用默认配置: {str(e)}")
        return DEFAULT_APP_CONFIG

# 加载应用配置
APP_CONFIG = load_app_config()

# 使用加载的应用配置初始化数据库超时配置
DB_TIMEOUT_CONFIG = get_db_timeout_config()

# 安全配置：请求频率限制
REQUEST_COUNTS = defaultdict(list)  # 存储每个IP的请求时间戳
MAX_REQUESTS_PER_MINUTE = 100  # 每分钟最大请求数
MAX_LOGIN_ATTEMPTS = 5  # 最大登录尝试次数
LOCKOUT_DURATION = 300  # 锁定持续时间（秒）
FAILED_LOGIN_ATTEMPTS = {}  # 存储失败的登录尝试
LOCKED_OUT_USERS = {}  # 存储被锁定的用户

def rate_limit_exceeded(ip_address):
    """检查是否超过请求频率限制"""
    now = time.time()
    # 清理一分钟前的请求记录
    REQUEST_COUNTS[ip_address] = [
        timestamp for timestamp in REQUEST_COUNTS[ip_address] 
        if now - timestamp < 60
    ]
    
    # 检查是否超过限制
    if len(REQUEST_COUNTS[ip_address]) >= MAX_REQUESTS_PER_MINUTE:
        return True
    
    # 记录当前请求
    REQUEST_COUNTS[ip_address].append(now)
    return False

def is_user_locked_out(username):
    """检查用户是否被锁定"""
    if username in LOCKED_OUT_USERS:
        if time.time() - LOCKED_OUT_USERS[username] < LOCKOUT_DURATION:
            return True
        else:
            # 锁定时间已过，清除锁定状态
            del LOCKED_OUT_USERS[username]
            if username in FAILED_LOGIN_ATTEMPTS:
                del FAILED_LOGIN_ATTEMPTS[username]
    return False

def record_failed_login_attempt(username):
    """记录失败的登录尝试"""
    now = time.time()
    if username not in FAILED_LOGIN_ATTEMPTS:
        FAILED_LOGIN_ATTEMPTS[username] = []
    
    # 清理5分钟前的尝试记录
    FAILED_LOGIN_ATTEMPTS[username] = [
        timestamp for timestamp in FAILED_LOGIN_ATTEMPTS[username]
        if now - timestamp < 300
    ]
    
    # 添加当前尝试
    FAILED_LOGIN_ATTEMPTS[username].append(now)
    
    # 检查是否达到最大尝试次数
    if len(FAILED_LOGIN_ATTEMPTS[username]) >= MAX_LOGIN_ATTEMPTS:
        LOCKED_OUT_USERS[username] = now  # 锁定用户

def validate_input(data, field_name, max_length=None, allowed_patterns=None):
    """验证输入数据"""
    if data is None:
        return False, f"{field_name} 不能为空"
    
    if not isinstance(data, str):
        data = str(data)
    
    # 检查长度
    if max_length and len(data) > max_length:
        return False, f"{field_name} 长度不能超过 {max_length} 个字符"
    
    # 根据字段名称决定验证策略
    if field_name.lower() == 'sql语句':
        # 对SQL语句使用较宽松的验证，主要在execute_sql中进行专门的安全检查
        # 这里只检查最基本的恶意模式，允许正常的SQL操作
        dangerous_patterns = [
            r"(?i)\b(shutdown|backup\s+database|restore\s+database)\b",
            r"(?i)(exec\s*\(|execute\s+\(|sp_|xp_)\s*[^']*;",
        ]
    else:
        # 对其他字段使用更严格的验证
        dangerous_patterns = [
            r"(?i)(drop\s+database|create\s+database|alter\s+database)",
            r"(?i)(exec\s*\(|execute\s+\(|sp_|xp_)",
            r"(?i)(;--|;|\*\/|--|#|/\*)",
            r"(?i)(grant\s+\w+\s+to|revoke\s+\w+\s+from)",
            r"(?i)(shutdown|backup\s+database)"
        ]
    
    for pattern in dangerous_patterns:
        if re.search(pattern, data):
            return False, f"{field_name} 包含非法字符或SQL注入尝试"
    
    # 如果提供了允许的模式，检查是否符合
    if allowed_patterns:
        for pattern in allowed_patterns:
            if not re.match(pattern, data):
                return False, f"{field_name} 格式不符合要求"
    
    return True, "验证通过"


def is_strong_password(password):
    """
    验证密码强度
    要求：至少6位，包含大写字母、小写字母、数字和特殊字符
    """
    import re
    
    # 检查长度
    if len(password) < 6:
        return False
    
    # 检查是否包含大写字母
    if not re.search(r'[A-Z]', password):
        return False
    
    # 检查是否包含小写字母
    if not re.search(r'[a-z]', password):
        return False
    
    # 检查是否包含数字
    if not re.search(r'\d', password):
        return False
    
    # 检查是否包含特殊字符
    if not re.search(r'[!@#$%^&*(),.?":{}|<>]', password):
        return False
    
    return True


def require_auth(f):
    """装饰器：验证用户是否已通过身份验证"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # 检查请求频率限制
        ip_address = request.environ.get('REMOTE_ADDR')
        if rate_limit_exceeded(ip_address):
            return jsonify({"status": "error", "message": "请求过于频繁，请稍后再试"}), 429
        
        # 检查是否需要密码验证
        app_password = load_app_password()
        if app_password and app_password.strip():  # 如果设置了应用密码
            session_token = request.headers.get('X-Session-Token') or request.cookies.get('session_token')
            if not session_token:
                return jsonify({"status": "error", "message": "请先进行身份验证"}), 401
            
            # 验证会话令牌（简单验证：检查是否存在于有效会话中）
            expected_token = hashlib.sha256(app_password.encode()).hexdigest()
            if session_token != expected_token:
                return jsonify({"status": "error", "message": "身份验证已过期或无效"}), 401
        
        return f(*args, **kwargs)
    return decorated_function

# 存储用户查询结果（使用UUID标识，解决并发问题）
QUERY_RESULTS = {}
# 结果过期时间（1小时）
RESULT_EXPIRE_TIME = 3600

# 初始化日志（记录SQL执行、导出等操作）
logging.basicConfig(
    filename=os.path.join(PROJECT_ROOT, 'log', 'sql_query_logs.log'),
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8'
)

# 支持的CSV分隔符
SUPPORTED_CSV_SEPARATORS = {
    'comma': ',',
    'semicolon': ';',
    'tab': '\t',
    'pipe': '|',
    'space': ' '
}

# 支持的Excel表头颜色
SUPPORTED_EXCEL_COLORS = {
    '4472C4': '蓝色（默认）',
    '5B9BD5': '浅蓝',
    '70AD47': '绿色',
    'FFC000': '黄色',
    'ED7D31': '橙色'
}

# 支持的数据库类型
SUPPORTED_DATABASES = ['postgresql', 'mysql', 'oracle', 'kingbase', 'tidb', 'oceanbase', 'highgo', 'gauss', 'uxdb', 'vastbase', 'greatdb', 'dm', 'yashandb', 'shentong', 'gbase', 'vanward']

# 常用SQL配置文件
COMMON_SQL_FILE = os.path.join(PROJECT_ROOT, "conf", "common_sql.json")

# 数据库密码加密密钥（简单异或+Base64，仅用于本地配置加密）
SECRET_KEY = b'data_check_sql_tool_key'

# 应用访问密码相关配置

# 默认常用SQL列表（首次运行会写入 common_sql.json）
DEFAULT_COMMON_SQLS = [
    {
        "id": "",
        "title": "查看版本",
        "sql": """select version();"""
    }
]


def encrypt_password(plain_password: str) -> str:
    """对数据库密码进行简单加密存储（异或 + Base64）。"""
    if not plain_password:
        return ""
    data = plain_password.encode('utf-8')
    xored = bytes([b ^ SECRET_KEY[i % len(SECRET_KEY)] for i, b in enumerate(data)])
    return "enc:" + base64.b64encode(xored).decode('ascii')


def decrypt_password(enc_password: str) -> str:
    """解密数据库密码，兼容未加密的旧数据。"""
    if not enc_password:
        return ""
    if not enc_password.startswith("enc:"):
        # 旧版本明文存储，直接返回
        return enc_password
    enc = enc_password[4:]
    try:
        xored = base64.b64decode(enc.encode('ascii'))
        data = bytes([b ^ SECRET_KEY[i % len(SECRET_KEY)] for i, b in enumerate(xored)])
        return data.decode('utf-8')
    except Exception as e:
        logging.error(f"解密数据库密码失败：{e}")
        return ""


def encrypt_app_password(plain_password: str) -> str:
    """对应用访问密码进行加密存储（异或 + Base64）。"""
    if not plain_password:
        return ""
    data = plain_password.encode('utf-8')
    xored = bytes([b ^ SECRET_KEY[i % len(SECRET_KEY)] for i, b in enumerate(data)])
    return "enc:" + base64.b64encode(xored).decode('ascii')


def decrypt_app_password(enc_password: str) -> str:
    """解密应用访问密码。"""
    if not enc_password:
        return ""
    if not enc_password.startswith("enc:"):
        return enc_password  # 兼容旧版本
    enc = enc_password[4:]
    try:
        xored = base64.b64decode(enc.encode('ascii'))
        data = bytes([b ^ SECRET_KEY[i % len(SECRET_KEY)] for i, b in enumerate(xored)])
        return data.decode('utf-8')
    except Exception as e:
        logging.error(f"解密应用访问密码失败：{e}")
        return ""


def load_app_password():
    """加载应用访问密码配置。"""
    try:
        # 从APP_CONFIG加载密码
        encrypted_password = APP_CONFIG.get('app_password', '')
        return decrypt_app_password(encrypted_password)
    except Exception as e:
        logging.error(f"加载应用访问密码失败：{e}")
        return ""


def save_app_password(plain_password: str):
    """保存应用访问密码配置（加密存储）。"""
    try:
        encrypted_password = encrypt_app_password(plain_password)
        
        # 更新APP_CONFIG中的密码
        APP_CONFIG['app_password'] = encrypted_password
        
        # 保存到配置文件
        with open(APP_CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(APP_CONFIG, f, ensure_ascii=False, indent=4)
        
        logging.info("应用访问密码已保存")
        return True
    except Exception as e:
        logging.error(f"保存应用访问密码失败：{e}")
        return False


def load_common_sqls():
    """加载常用SQL列表（优先读取JSON文件，无则写入默认列表）。"""
    try:
        with open(COMMON_SQL_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if isinstance(data, list):
                return data
    except Exception as e:
        logging.warning(f"读取常用SQL配置失败，将使用默认配置：{e}")
    # 使用默认列表并尝试写入文件
    sql_list = DEFAULT_COMMON_SQLS
    # 为每个默认SQL生成唯一ID
    for item in sql_list:
        if not item.get("id"):
            item["id"] = str(uuid.uuid4())
    try:
        with open(COMMON_SQL_FILE, 'w', encoding='utf-8') as f:
            json.dump(sql_list, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logging.error(f"写入默认常用SQL配置失败：{e}")
    return sql_list


def save_common_sqls(sql_list):
    """保存常用SQL列表到JSON文件。"""
    try:
        with open(COMMON_SQL_FILE, 'w', encoding='utf-8') as f:
            json.dump(sql_list, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logging.error(f"保存常用SQL配置失败：{e}")

# ===================== 工具函数 =====================
def load_db_config():
    """加载数据库配置（优先读取本地文件，并对密码做加解密处理）"""
    try:
        with open(DB_CONFIG_FILE, 'r', encoding='utf-8') as f:
            config = json.load(f)
    except Exception:
        # 无配置文件则用默认，并生成配置文件（密码加密后存储）
        config = DEFAULT_DB_CONFIG.copy()
        if config.get('password'):
            config['password'] = encrypt_password(config['password'])
        try:
            with open(DB_CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logging.error(f"写入默认数据库配置失败：{e}")
    # 运行时始终返回明文密码，方便后续连接使用
    if 'password' in config and config['password']:
        config['password'] = decrypt_password(config['password'])
    return config


def load_multi_db_config():
    """加载多数据库配置（兼容旧格式）"""
    try:
        with open(DB_CONFIG_FILE, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        # 检查是否为新格式（多数据库配置）
        if 'databases' in config:
            # 解密所有数据库配置的密码
            for db_config in config['databases']:
                if db_config.get('password'):
                    db_config['password'] = decrypt_password(db_config['password'])
            return config
        else:
            # 旧格式，转换为新格式
            config['password'] = decrypt_password(config.get('password', ''))
            new_format = {
                "databases": [
                    {
                        "id": "legacy_db",
                        "name": "旧版数据库配置",
                        "type": config.get('type', 'postgresql'),
                        "host": config['host'],
                        "port": config['port'],
                        "user": config['user'],
                        "password": config['password'],
                        "database": config['database'],
                        "is_default": True
                    }
                ]
            }
            return new_format
    except Exception as e:
        logging.error(f"加载数据库配置失败：{str(e)}")
        # 返回默认配置
        return {
            "databases": [
                {
                    "id": "default_db",
                    "name": "默认数据库",
                    "type": "postgresql",
                    "host": "localhost",
                    "port": "5432",
                    "user": "postgres",
                    "password": "123456",
                    "database": "postgres",
                    "is_default": True
                }
            ]
        }


def get_database_by_id(db_id):
    """根据ID获取数据库配置"""
    config = load_multi_db_config()
    databases = config.get('databases', [])
    
    for db in databases:
        if db['id'] == db_id:
            return db
    
    return None


def get_default_database():
    """获取默认数据库配置"""
    config = load_multi_db_config()
    databases = config.get('databases', [])
    
    # 查找标记为默认的数据库
    for db in databases:
        if db.get('is_default', False):
            return db
    
    # 如果没有默认数据库，返回第一个
    if databases:
        return databases[0]
    
    return None

def get_db_connection(db_id=None):
    """获取数据库连接上下文管理器（支持多数据库类型，确保使用正确的驱动）"""
    db_config = None
    if db_id:
        db_config = get_database_by_id(db_id)
    else:
        db_config = get_default_database()
    
    if not db_config:
        raise ValueError("未找到有效的数据库配置")
    
    db_type = db_config.get('type', 'postgresql').lower()
    
    # 数据库驱动映射表 - 确保每种数据库使用正确的驱动
    DRIVER_MAPPING = {
        # PostgreSQL系列（包括兼容PostgreSQL协议的国产数据库）
        'postgresql': ('PostgreSQL', get_postgresql_connection, 'psycopg2'),
        'highgo': ('瀚高数据库', get_postgresql_connection, 'psycopg2'),
        'gauss': ('华为高斯数据库', get_postgresql_connection, 'psycopg2'),
        'uxdb': ('优图数据库', get_postgresql_connection, 'psycopg2'),
        'vastbase': ('海量数据库', get_postgresql_connection, 'psycopg2'),
        'yashandb': ('崖山数据库', get_yashandb_connection, 'yasdb'),
        'gbase': ('南大通用数据库', get_postgresql_connection, 'psycopg2'),
        'vanward': ('万里数据库', get_postgresql_connection, 'psycopg2'),
        'kingbase': ('人大金仓数据库', get_kingbase_connection, 'psycopg2'),
        
        # MySQL系列（包括兼容MySQL协议的数据库）
        'mysql': ('MySQL', get_mysql_connection, 'pymysql'),
        'tidb': ('TiDB分布式数据库', get_tidb_connection, 'pymysql'),
        'oceanbase': ('OceanBase数据库', get_oceanbase_connection, 'pymysql'),
        'greatdb': ('巨杉数据库', get_mysql_connection, 'pymysql'),
        
        # Oracle系列（包括兼容Oracle协议的数据库）
        'oracle': ('Oracle数据库', get_oracle_connection, 'cx_Oracle'),
        'shentong': ('神通数据库', get_oracle_connection, 'cx_Oracle'),
        
        # 专用驱动数据库
        'dm': ('达梦数据库', get_dm_connection, 'dm_python')
    }
    
    # 检查数据库类型是否支持
    if db_type not in DRIVER_MAPPING:
        supported_types = ', '.join(DRIVER_MAPPING.keys())
        raise ValueError(f"不支持的数据库类型：{db_type}。支持的类型：{supported_types}")
    
    # 获取驱动信息
    db_display_name, connection_func, required_driver = DRIVER_MAPPING[db_type]
    
    # 记录驱动使用日志
    logging.info(f"数据库连接 [{db_config['name']}] 使用驱动: {db_display_name} ({required_driver})")
    
    # 返回连接函数和配置
    return connection_func, db_config


def execute_db_operation(db_id, operation_func):
    """执行数据库操作的通用函数"""
    conn_manager, db_config = get_db_connection(db_id)
    with conn_manager(db_config) as conn:
        return operation_func(conn)


def set_default_database(db_id):
    """设置默认数据库"""
    config = load_multi_db_config()
    databases = config.get('databases', [])
    
    # 重置所有数据库的默认标志
    for db in databases:
        db['is_default'] = False
    
    # 设置指定数据库为默认
    for db in databases:
        if db['id'] == db_id:
            db['is_default'] = True
            break
    
    # 保存配置
    for db_config in config['databases']:
        raw_password = db_config.get('password', '')
        if raw_password:
            db_config['password'] = encrypt_password(raw_password)
    
    with open(DB_CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    
    return True



@contextmanager
def get_yashandb_connection(db_config):
    """崖山数据库连接上下文管理器（使用专用yasdb驱动）"""
    conn = None
    try:
        import yasdb
        # 解密密码
        decrypted_password = decrypt_password(db_config.get('password', ''))
        
        # 构造DSN连接字符串
        dsn = f"{db_config['host']}:{db_config['port']}"
        
        # 使用DSN方式连接
        conn = yasdb.connect(
            dsn=dsn,
            user=db_config['user'],
            password=decrypted_password,
        )
        yield conn
    except ImportError:
        raise Exception("未找到崖山数据库驱动(yasdb)，请安装yasdb包")
    except Exception as e:
        if conn:
            conn.rollback()
        raise Exception(f"崖山数据库连接失败: {str(e)}")
    finally:
        if conn:
            conn.close()


def get_postgresql_connection(db_config):
    """PostgreSQL数据库连接上下文管理器"""
    conn = None
    try:
        import psycopg2
        from psycopg2 import OperationalError, ProgrammingError
        # 解密密码
        decrypted_password = decrypt_password(db_config.get('password', ''))
        conn = psycopg2.connect(
            host=db_config['host'],
            port=int(db_config['port']),  # 确保端口是整数
            user=db_config['user'],
            password=decrypted_password,
            database=db_config['database'],
            connect_timeout=10  # 连接超时时间
        )
        yield conn
    except Exception as e:
        if conn:
            conn.rollback()
        raise e
    finally:
        if conn:
            conn.close()



# 为其他数据库类型预留连接函数
@contextmanager
def get_mysql_connection(db_config):
    """MySQL数据库连接上下文管理器"""
    try:
        import pymysql
        # 解密密码
        decrypted_password = decrypt_password(db_config.get('password', ''))
        conn = pymysql.connect(
            host=db_config['host'],
            port=int(db_config['port']),
            user=db_config['user'],
            password=decrypted_password,
            database=db_config['database'],
            charset='utf8mb4',
            connect_timeout=10
        )
        yield conn
    except ImportError:
        raise ImportError("请安装PyMySQL：pip install PyMySQL")
    except Exception as e:
        raise e
    finally:
        if 'conn' in locals() and conn:
            conn.close()



@contextmanager
def get_oracle_connection(db_config):
    """Oracle数据库连接上下文管理器"""
    try:
        import cx_Oracle
        # 解密密码
        decrypted_password = decrypt_password(db_config.get('password', ''))
        dsn = cx_Oracle.makedsn(db_config['host'], int(db_config['port']), service_name=db_config['database'])
        conn = cx_Oracle.connect(
            user=db_config['user'],
            password=decrypted_password,
            dsn=dsn,
            encoding="UTF-8"
        )
        yield conn
    except ImportError:
        raise ImportError("请安装cx_Oracle：pip install cx_Oracle")
    except Exception as e:
        raise e
    finally:
        if 'conn' in locals() and conn:
            conn.close()



@contextmanager
def get_kingbase_connection(db_config):
    """人大金仓数据库连接上下文管理器"""
    # 金仓数据库通常兼容PostgreSQL协议，所以可以使用psycopg2
    try:
        import psycopg2 as kingbase_psycopg2
        # 解密密码
        decrypted_password = decrypt_password(db_config.get('password', ''))
        conn = kingbase_psycopg2.connect(
            host=db_config['host'],
            port=int(db_config['port']),
            user=db_config['user'],
            password=decrypted_password,
            database=db_config['database'],
            connect_timeout=10
        )
        yield conn
    except ImportError:
        raise ImportError("请安装psycopg2：pip install psycopg2")
    except Exception as e:
        raise e
    finally:
        if 'conn' in locals() and conn:
            conn.close()



@contextmanager
def get_tidb_connection(db_config):
    """TiDB数据库连接上下文管理器"""
    # TiDB兼容MySQL协议，所以使用PyMySQL
    return get_mysql_connection(db_config)



@contextmanager
def get_oceanbase_connection(db_config):
    """OceanBase数据库连接上下文管理器"""
    # OceanBase兼容MySQL协议，所以使用PyMySQL
    return get_mysql_connection(db_config)


@contextmanager
def get_dm_connection(db_config):
    """达梦数据库连接上下文管理器"""
    try:
        import dm_python as dm
        # 解密密码
        decrypted_password = decrypt_password(db_config.get('password', ''))
        conn = dm.connect(
            host=db_config['host'],
            port=int(db_config['port']),
            user=db_config['user'],
            password=decrypted_password,
            database=db_config['database'],
            connect_timeout=10
        )
        yield conn
    except ImportError:
        raise ImportError("请安装达梦数据库驱动：pip install dm-python")
    except Exception as e:
        raise e
    finally:
        if 'conn' in locals() and conn:
            conn.close()


def split_sql_statements(sql_content):
    """分割SQL语句，支持多种分隔符"""
    import re
    # 使用正则表达式分割SQL语句，支持 ; 或 GO 作为分隔符
    # 先移除多行注释 /* */ 和单行注释 --
    cleaned_sql = re.sub(r'/\*.*?\*/', '', sql_content, flags=re.DOTALL)
    cleaned_sql = re.sub(r'--.*?$', '', cleaned_sql, flags=re.MULTILINE)
    
    # 按分号分割，并去除空语句
    statements = [stmt.strip() for stmt in cleaned_sql.split(';') if stmt.strip()]
    return statements

def check_sql_safety(sql):
    """增强版SQL安全校验：禁止危险操作，但允许合法的数据操作"""
    # 移除注释和空格，避免绕过检测
    clean_sql = re.sub(r'/\*.*?\*/|--.*?$', '', sql, flags=re.DOTALL | re.MULTILINE).strip().upper()
    clean_sql = re.sub(r'\s+', ' ', clean_sql)
    
    # 危险操作关键词（更全面）
    dangerous_keywords = [
        'DROP', 'TRUNCATE', 'ALTER', 'CREATE', 'RENAME', 'GRANT', 'REVOKE', 'LOCK', 'SHUTDOWN'
    ]
    
    # 检查危险关键词（独立单词匹配，避免误判）
    for keyword in dangerous_keywords:
        # 检查是否是独立的关键词，而不是其他词的一部分（如comment中的关键词）
        if re.search(rf'\b{keyword}\b', clean_sql):
            return False, f"禁止执行包含「{keyword}」的危险SQL！"
    
    # 检查可能的SQL注入模式
    injection_patterns = [
        r"(?i)(\bUNION\b.*\bSELECT\b)",  # UNION注入
        r"(?i)(\bOR\b\s+\d+\s*=\s*\d+)",  # 永真式注入
        r"(?i)(exec\s*\(|execute\s+\(|sp_|xp_)\s*[^']*;",  # 存储过程执行
    ]
    
    for pattern in injection_patterns:
        if re.search(pattern, sql):
            return False, "检测到潜在的SQL注入攻击模式！"
    
    return True, ""

def clean_expired_data():
    """清理过期的查询结果和临时文件"""
    try:
        # 清理过期查询结果
        current_time = time.time()
        expired_keys = []
        for key, value in QUERY_RESULTS.items():
            if current_time - value['create_time'] > RESULT_EXPIRE_TIME:
                expired_keys.append(key)
        for key in expired_keys:
            del QUERY_RESULTS[key]
            logging.info(f"清理过期查询结果：{key}")
        
        # 清理2小时前的临时文件（Excel和CSV）
        for file in os.listdir(TEMP_DIR):
            if (file.startswith('SQL查询结果_') and 
                (file.endswith('.xlsx') or file.endswith('.csv'))):
                file_path = os.path.join(TEMP_DIR, file)
                if current_time - os.path.getctime(file_path) > 7200:  # 2小时
                    os.remove(file_path)
                    logging.info(f"清理过期文件：{file}")
    except Exception as e:
        logging.error(f"清理过期数据失败：{str(e)}")

def set_excel_style(ws, header_color="4472C4"):
    """设置Excel样式（支持自定义表头颜色）"""
    # 字体配置
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    content_font = Font(name='微软雅黑', size=10)
    # 对齐方式
    align = Alignment(horizontal='center', vertical='center')
    # 边框样式
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # 表头背景色
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
    
    # 应用样式
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = align
        cell.border = border
        cell.fill = header_fill
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = content_font
            cell.alignment = align
            cell.border = border
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def generate_csv_content(columns, results, separator=',', include_header=True):
    """生成CSV内容"""
    # 使用StringIO处理CSV内容
    output = StringIO()
    # 设置CSV写入器，处理中文和特殊字符
    writer = csv.writer(output, delimiter=separator, 
                       quoting=csv.QUOTE_MINIMAL,
                       lineterminator='\n',
                       escapechar='\\')
    
    # 写入表头
    if include_header and columns:
        writer.writerow(columns)
    
    # 写入数据行
    for row in results:
        # 处理None值和特殊类型
        cleaned_row = []
        for cell in row:
            if cell is None:
                cleaned_row.append('')
            elif isinstance(cell, (datetime,)):
                cleaned_row.append(cell.strftime('%Y-%m-%d %H:%M:%S'))
            else:
                cleaned_row.append(str(cell))
        writer.writerow(cleaned_row)
    
    # 重置指针到开头
    output.seek(0)
    # 转换为BytesIO并编码为UTF-8（带BOM，解决Excel打开中文乱码）
    csv_bytes = BytesIO()
    # 写入BOM头
    csv_bytes.write(b'\xef\xbb\xbf')
    # 写入CSV内容
    csv_bytes.write(output.getvalue().encode('utf-8'))
    csv_bytes.seek(0)
    
    return csv_bytes

# ===================== 路由 =====================
@app.route('/')
def index():
    """首页"""
    # 每次访问首页清理一次过期数据
    clean_expired_data()
    # 传递数据库配置和列表到前端
    config = load_multi_db_config()
    return render_template('index.html', 
                           db_config=config,
                           databases=config.get('databases', []),
                           excel_colors=SUPPORTED_EXCEL_COLORS,
                           supported_dbs=SUPPORTED_DATABASES,
                           has_password=(load_app_password() != ""))

@app.route('/sql_beautify_test')
def sql_beautify_test():
    """SQL美化功能测试页面"""
    return render_template('sql_beautify_test.html')

@app.route('/save_db_config', methods=['POST'])
@require_auth
def save_db_config():
    """保存数据库配置（支持多数据库配置）"""
    with DB_CONFIG_LOCK:  # 添加锁
        try:
            config = request.json
            # 添加调试日志
            logging.info(f"DEBUG: Received config keys: {list(config.keys()) if isinstance(config, dict) else type(config)}")
            if isinstance(config, dict) and 'databases' in config:
                logging.info(f"DEBUG: Detected multi-db config with {len(config.get('databases', []))} databases")
            else:
                logging.info(f"DEBUG: Detected single-db config with id: {config.get('id', 'NO_ID')}")
            
            # 检查是否是多数据库配置格式
            if 'databases' in config:
                # 是多数据库配置，直接保存
                # 智能处理数据库配置的密码：如果密码已加密（以"enc:"开头）则跳过，否则加密
                for db_config in config['databases']:
                    raw_password = db_config.get('password', '')
                    if raw_password:
                        # 检查密码是否已经加密
                        if not raw_password.startswith("enc:"):
                            # 如果没有加密，则进行加密
                            db_config['password'] = encrypt_password(raw_password)
                        else:
                            # 如果已经加密，则保持原样
                            db_config['password'] = raw_password
                
                with open(DB_CONFIG_FILE, 'w', encoding='utf-8') as f:
                    json.dump(config, f, ensure_ascii=False, indent=2)
                
                logging.info("多数据库配置已更新")
                return jsonify({"status": "success", "message": "多数据库配置保存成功！"})
            else:
                # 旧格式，转换为新格式
                # 验证必要字段
                required_fields = ['host', 'port', 'user', 'database', 'type']
                for field in required_fields:
                    if not config.get(field):
                        return jsonify({"status": "error", "message": f"{field}为必填项！"})
                
                # 严格校验端口格式
                if not config['port'].isdigit():
                    return jsonify({"status": "error", "message": "端口必须为数字！"})
                
                # 检查数据库类型是否支持
                db_type = config.get('type', 'postgresql').lower()
                if db_type not in SUPPORTED_DATABASES:
                    return jsonify({"status": "error", "message": f"不支持的数据库类型：{db_type}"})
                
                # 读取现有配置
                try:
                    with open(DB_CONFIG_FILE, 'r', encoding='utf-8') as f:
                        existing_config = json.load(f)
                except FileNotFoundError:
                    existing_config = {"databases": []}
                except json.JSONDecodeError:
                    existing_config = {"databases": []}
                
                # 生成新配置
                raw_password = config.get('password', '')
                # 如果ID为'None'或None，则生成新的UUID
                config_id = config.get('id')
                if config_id is None or config_id == 'None':
                    config_id = str(uuid.uuid4())
                
                # 智能处理密码：如果密码已加密（以"enc:"开头）则跳过，否则加密
                processed_password = ""
                if raw_password:
                    raw_password = raw_password.strip()
                    if not raw_password.startswith("enc:"):
                        # 如果没有加密，则进行加密
                        processed_password = encrypt_password(raw_password)
                    else:
                        # 如果已经加密，则保持原样
                        processed_password = raw_password
                
                new_db_config = {
                    "id": config_id,
                    "name": config.get('name', f"{db_type.capitalize()}数据库"),
                    "type": db_type,
                    "host": config['host'].strip(),
                    "port": config['port'].strip(),
                    "user": config['user'].strip(),
                    "password": processed_password,
                    "database": config['database'].strip(),
                    "is_default": config.get('is_default', False)
                }
                
                # 保存前备份现有数据库的密码字段，以防意外修改
                existing_passwords = {}
                for db in existing_config.get('databases', []):
                    if 'password' in db and db['password']:
                        existing_passwords[db['id']] = db['password']
                            
                # 检查是否已存在相同ID的数据库配置，如果存在则更新，否则添加
                db_exists = False
                for i, db in enumerate(existing_config.get('databases', [])):
                    if db['id'] == new_db_config['id']:
                        existing_config['databases'][i] = new_db_config
                        db_exists = True
                        break
                            
                if not db_exists:
                    existing_config.setdefault('databases', []).append(new_db_config)
                            
                # 恢复现有数据库的密码字段（如果被意外修改）
                for db in existing_config.get('databases', []):
                    if db['id'] in existing_passwords and db.get('password') != existing_passwords[db['id']]:
                        # 只有当密码为空或与备份不同时才恢复
                        if not db.get('password'):
                            db['password'] = existing_passwords[db['id']]
                
                with open(DB_CONFIG_FILE, 'w', encoding='utf-8') as f:
                    json.dump(existing_config, f, ensure_ascii=False, indent=2)
                        
                logging.info(f"数据库配置已更新：{db_type}://{config['host']}:{config['port']}/{config['database']}")
                return jsonify({"status": "success", "message": "配置保存成功！"})
        
        except Exception as e:
            logging.error(f"保存数据库配置失败：{str(e)}")
            return jsonify({"status": "error", "message": f"保存失败：{str(e)}"})

@app.route('/test_db_connection', methods=['POST'])
@require_auth
def test_db_connection():
    """测试数据库连接（支持多数据库类型，显示使用的驱动信息）"""
    try:
        config = request.json
        # 基础校验
        if not all([config.get('host'), config.get('port'), config.get('database'), config.get('type'), config.get('user')]):
            return jsonify({"status": "error", "message": "缺少必要的数据库配置项！"})
        
        db_type = config.get('type', 'postgresql').lower()
        
        # 数据库驱动映射表
        DRIVER_INFO = {
            'postgresql': 'PostgreSQL (psycopg2)',
            'mysql': 'MySQL (PyMySQL)',
            'oracle': 'Oracle (cx_Oracle)',
            'kingbase': '人大金仓 (psycopg2)',
            'tidb': 'TiDB (PyMySQL)',
            'oceanbase': 'OceanBase (PyMySQL)',
            'highgo': '瀚高数据库 (psycopg2)',
            'gauss': '华为高斯数据库 (psycopg2)',
            'uxdb': '优图数据库 (psycopg2)',
            'vastbase': '海量数据库 (psycopg2)',
            'yashandb': '崖山数据库 (yasdb)',
            'shentong': '神通数据库 (cx_Oracle)',
            'greatdb': '巨杉数据库 (PyMySQL)',
            'dm': '达梦数据库 (dm_python)',
            'gbase': '南大通用数据库 (psycopg2)',
            'vanward': '万里数据库 (psycopg2)'
        }
        
        driver_info = DRIVER_INFO.get(db_type, f'未知驱动 ({db_type})')
        logging.info(f"测试数据库连接 [{config.get('database')}] 使用驱动: {driver_info}")
        
        if db_type == 'postgresql':
            # 修复：connect_timeout 作为独立参数传入，而非options
            # 注意：port需要转为整数，connect_timeout单位是秒
            import psycopg2
            # 解密密码
            decrypted_password = decrypt_password(config.get('password', ''))
            conn = psycopg2.connect(
                host=config['host'],
                port=int(config['port']),  # 关键：端口必须是整数
                user=config['user'],
                password=decrypted_password,
                database=config['database'],
                connect_timeout=5  # 独立参数设置5秒连接超时
            )
            conn.close()
        elif db_type == 'mysql':
            import pymysql
            # 解密密码
            decrypted_password = decrypt_password(config.get('password', ''))
            conn = pymysql.connect(
                host=config['host'],
                port=int(config['port']),
                user=config['user'],
                password=decrypted_password,
                database=config['database'],
                charset='utf8mb4',
                connect_timeout=5
            )
            conn.close()
        elif db_type == 'oracle':
            import cx_Oracle
            # 解密密码
            decrypted_password = decrypt_password(config.get('password', ''))
            dsn = cx_Oracle.makedsn(config['host'], int(config['port']), service_name=config['database'])
            conn = cx_Oracle.connect(
                user=config['user'],
                password=decrypted_password,
                dsn=dsn,
                encoding="UTF-8",
                timeout=5
            )
            conn.close()
        elif db_type == 'kingbase':  # 人大金仓
            import psycopg2 as kingbase_psycopg2
            # 解密密码
            decrypted_password = decrypt_password(config.get('password', ''))
            conn = kingbase_psycopg2.connect(
                host=config['host'],
                port=int(config['port']),
                user=config['user'],
                password=decrypted_password,
                database=config['database'],
                connect_timeout=5
            )
            conn.close()
        elif db_type in ['tidb', 'oceanbase', 'greatdb']:  # TiDB、OceanBase、GreatDB兼容MySQL协议
            import pymysql
            # 解密密码
            decrypted_password = decrypt_password(config.get('password', ''))
            conn = pymysql.connect(
                host=config['host'],
                port=int(config['port']),
                user=config['user'],
                password=decrypted_password,
                database=config['database'],
                charset='utf8mb4',
                connect_timeout=5
            )
            conn.close()
        elif db_type in ['highgo', 'gauss', 'uxdb', 'vastbase', 'gbase', 'vanward']:  # PostgreSQL兼容的国产数据库
            import psycopg2
            # 解密密码
            decrypted_password = decrypt_password(config.get('password', ''))
            conn = psycopg2.connect(
                host=config['host'],
                port=int(config['port']),
                user=config['user'],
                password=decrypted_password,
                database=config['database'],
                connect_timeout=5
            )
            conn.close()
        elif db_type == 'yashandb':  # 崖山数据库使用专用yasdb驱动
            import yasdb
            # 解密密码
            decrypted_password = decrypt_password(config.get('password', ''))
            # 构造DSN连接字符串
            dsn = f"{config['host']}:{config['port']}"
            conn = yasdb.connect(
                dsn=dsn,
                user=config['user'],
                password=decrypted_password,
            )
            conn.close()
        elif db_type == 'shentong':  # 神通数据库兼容Oracle协议
            import cx_Oracle
            # 解密密码
            decrypted_password = decrypt_password(config.get('password', ''))
            dsn = cx_Oracle.makedsn(config['host'], int(config['port']), service_name=config['database'])
            conn = cx_Oracle.connect(
                user=config['user'],
                password=decrypted_password,
                dsn=dsn,
                encoding="UTF-8",
                timeout=5
            )
            conn.close()
        elif db_type == 'dm':  # 达梦数据库
            import dm_python as dm
            # 解密密码
            decrypted_password = decrypt_password(config.get('password', ''))
            conn = dm.connect(
                host=config['host'],
                port=int(config['port']),
                user=config['user'],
                password=decrypted_password,
                database=config['database'],
                connect_timeout=5
            )
            conn.close()
        else:
            return jsonify({"status": "error", "message": f"不支持的数据库类型：{db_type}"})
        
        logging.info(f"数据库连接测试成功：{db_type}://{config['host']}:{config['port']}/{config['database']} (使用 {driver_info})")
        return jsonify({"status": "success", "message": f"数据库连接成功！使用驱动: {driver_info}"})
    
    except ValueError as e:
        # 端口转换整数失败
        logging.error(f"数据库端口格式错误：{str(e)}")
        return jsonify({"status": "error", "message": f"端口格式错误：{str(e)}"})
    except ImportError as e:
        logging.error(f"缺少数据库驱动：{str(e)}")
        missing_package = str(e).split()[-1] if len(str(e).split()) > 0 else "unknown"
        return jsonify({"status": "error", "message": f"缺少数据库驱动，请安装：{missing_package}"})
    except Exception as e:
        logging.error(f"数据库连接测试失败：{str(e)}")
        return jsonify({"status": "error", "message": f"连接失败：{str(e)}"})

@app.route('/common_sqls', methods=['GET', 'POST', 'DELETE'])
def common_sqls():
    """常用SQL管理接口：GET读取，POST新增/更新，DELETE删除。"""
    try:
        if request.method == 'GET':
            sql_list = load_common_sqls()
            return jsonify({"status": "success", "data": sql_list})
        elif request.method == 'POST':
            data = request.json or {}
            sql_id = data.get('id')
            title = (data.get('title') or '').strip()
            sql_text = (data.get('sql') or '').strip()
            
            if not title or not sql_text:
                return jsonify({"status": "error", "message": "标题和SQL内容不能为空！"})
            
            sql_list = load_common_sqls()
            if sql_id:
                # 更新
                found = False
                for item in sql_list:
                    if item.get('id') == sql_id:
                        item['title'] = title
                        item['sql'] = sql_text
                        found = True
                        break
                if not found:
                    return jsonify({"status": "error", "message": "未找到对应的SQL记录！"})
                msg = "更新成功！"
            else:
                # 新增
                new_item = {
                    "id": str(uuid.uuid4()),
                    "title": title,
                    "sql": sql_text
                }
                sql_list.append(new_item)
                msg = "保存成功！"
            
            save_common_sqls(sql_list)
            return jsonify({"status": "success", "message": msg})
        
        elif request.method == 'DELETE':
            sql_id = request.args.get('id')
            if not sql_id:
                return jsonify({"status": "error", "message": "缺少ID参数！"})
            
            sql_list = load_common_sqls()
            new_list = [item for item in sql_list if item.get('id') != sql_id]
            if len(new_list) == len(sql_list):
                return jsonify({"status": "error", "message": "未找到对应的SQL记录！"})
            
            save_common_sqls(new_list)
            return jsonify({"status": "success", "message": "删除成功！"})
            
    except Exception as e:
        logging.error(f"常用SQL接口执行失败：{e}")
        return jsonify({"status": "error", "message": f"操作失败：{str(e)}"})


@app.route('/import_common_sqls', methods=['POST'])
def import_common_sqls():
    """导入常用SQL列表：覆盖现有列表。"""
    try:
        data = request.json
        if not isinstance(data, list):
            return jsonify({"status": "error", "message": "格式错误，必须为数组！"})
        
        # 为没有ID的项目生成ID
        for item in data:
            if not item.get('id'):
                item['id'] = str(uuid.uuid4())
            
        save_common_sqls(data)
        logging.info(f"成功导入 {len(data)} 条常用SQL")
        return jsonify({"status": "success", "message": f"成功导入 {len(data)} 条常用SQL"})
    except Exception as e:
        logging.error(f"导入常用SQL失败：{e}")
        return jsonify({"status": "error", "message": str(e)})


@app.route('/export_common_sqls')
def export_common_sqls():
    """导出常用SQL列表为JSON文件（支持自定义文件名）。"""
    try:
        custom_filename = request.args.get('filename')
        if not custom_filename:
            custom_filename = f"common_sqls_{datetime.now().strftime('%Y%m%d')}.json"
        
        if not os.path.exists(COMMON_SQL_FILE):
            return jsonify({"status": "error", "message": "常用SQL文件不存在！"})
        
        return send_file(
            os.path.abspath(COMMON_SQL_FILE),
            as_attachment=True,
            download_name=custom_filename,
            mimetype='application/json'
        )
    except Exception as e:
        logging.error(f"导出常用SQL失败：{e}")
        return jsonify({"status": "error", "message": str(e)})


@app.route('/execute_sql', methods=['POST'])
@require_auth
def execute_sql():
    """执行SQL（支持单条或多条语句，优化并发和内存使用，支持多数据库）"""
    try:
        # 获取参数
        sql = request.form.get('sql', '').strip()
        page = int(request.form.get('page', 1))
        page_size = int(request.form.get('page_size', 50))
        db_id = request.form.get('db_id', None)  # 新增：数据库ID
        
        # 输入验证
        is_valid, message = validate_input(sql, "SQL语句", max_length=10000)
        if not is_valid:
            return jsonify({"status": "error", "message": message})
        
        # 额外验证SQL语句的安全性 - 仅检查明确的危险操作
        # 注意：SELECT, INSERT, UPDATE, DELETE 是正常操作，不应阻止
        dangerous_keywords = [
            # 数据库/表结构删除操作
            r"(?i)\b(drop\s+table|truncate\s+table)\b",
            r"(?i)\b(drop\s+(?!index|procedure|view|trigger|function)\w+)\b",  # 排除合法的drop操作如drop index
            r"(?i)\b(shutdown|backup\s+database|restore\s+database)\b",
            # 系统级危险操作
            r"(?i)(exec\s*\(|execute\s+\(|sp_|xp_)[^']*;",  # exec等危险函数，但允许在字符串中出现
            # 权限管理操作
            r"(?i)\b(grant\s+\w+\s+to|revoke\s+\w+\s+from)\b",
        ]
        
        for pattern in dangerous_keywords:
            if re.search(pattern, sql):
                logging.warning(f"检测到潜在危险SQL操作: {sql[:100]}...")
                return jsonify({"status": "error", "message": "SQL语句包含被禁止的操作！"})
        
        # 参数校验
        if not sql:
            return jsonify({"status": "error", "message": "请输入SQL语句！"})
        if page < 1 or page_size < 1 or page_size > 1000:
            return jsonify({"status": "error", "message": "页码需≥1，每页条数需1-1000之间！"})
        
        # 分割SQL语句
        sql_statements = split_sql_statements(sql)
        
        if len(sql_statements) == 1:
            # 单条语句执行（原有逻辑）
            return execute_single_statement(sql_statements[0], page, page_size, db_id)
        else:
            # 多条语句执行
            results = []
            for i, stmt in enumerate(sql_statements):
                stmt = stmt.strip()
                if stmt:  # 忽略空语句
                    try:
                        result = execute_single_statement(stmt, page, page_size, db_id)
                        if result.get('status') == 'success':
                            result['statement_index'] = i + 1
                            result['original_sql'] = stmt[:100] + "..." if len(stmt) > 100 else stmt
                            results.append(result)
                        else:
                            result['statement_index'] = i + 1
                            result['original_sql'] = stmt[:100] + "..." if len(stmt) > 100 else stmt
                            results.append(result)
                            # 如果是查询语句出错，继续执行下一条
                            if 'SELECT' in stmt.upper() or 'EXPLAIN' in stmt.upper():
                                continue
                            else:
                                break  # 非查询语句出错则停止
                    except Exception as e:
                        results.append({
                            "status": "error",
                            "message": f"第{i+1}条语句执行失败: {str(e)}",
                            "original_sql": stmt[:100] + "..." if len(stmt) > 100 else stmt,
                            "statement_index": i + 1
                        })
            
            return jsonify({
                "status": "success",
                "message": f"共执行{len(results)}条语句",
                "results": results,
                "is_batch": True
            })
    
    except ValueError as e:
        logging.error(f"参数转换失败：{str(e)}")
        return jsonify({"status": "error", "message": f"参数格式错误：{str(e)}"})
    except Exception as e:
        logging.error(f"SQL执行失败：{str(e)} | SQL：{sql[:100]}...")
        return jsonify({"status": "error", "message": f"执行失败：{str(e)}"})


@app.route('/analyze_query_plan', methods=['POST'])
@require_auth
def analyze_query_plan():
    """分析SQL查询计划（支持多数据库类型）"""
    try:
        # 获取参数
        sql = request.form.get('sql', '').strip()
        db_id = request.form.get('db_id', None)
        
        # 输入验证
        is_valid, message = validate_input(sql, "SQL语句", max_length=5000)
        if not is_valid:
            return jsonify({"status": "error", "message": message})
            
        if not sql:
            return jsonify({"status": "error", "message": "请输入SQL语句！"})
            
        # 验证SQL是否为SELECT语句 - 改进的验证逻辑
        # 移除注释后再检查是否以SELECT开头
        clean_sql = re.sub(r'/\*.*?\*/|--.*?$', '', sql, flags=re.DOTALL | re.MULTILINE).strip()
        clean_sql_upper = clean_sql.upper()
        
        # 支持多种SELECT语句格式
        is_select_query = (
            clean_sql_upper.startswith('SELECT') or 
            clean_sql_upper.startswith('EXPLAIN') or
            (clean_sql_upper.startswith('WITH') and 'SELECT' in clean_sql_upper)
        )
        
        if not is_select_query:
            return jsonify({"status": "error", "message": "查询计划分析仅支持SELECT/EXPLAIN/WITH查询语句！"})
            
        # 获取数据库连接
        get_connection_func, db_config = get_db_connection(db_id)
        db_type = db_config.get('type', 'postgresql').lower()
        
        # 构建EXPLAIN语句
        explain_sql = ""
        if db_type in ['postgresql', 'kingbase', 'highgo', 'gauss', 'uxdb', 'vastbase', 'yashandb', 'gbase', 'vanward']:
            explain_sql = f"EXPLAIN ANALYZE {sql}"
        elif db_type in ['mysql', 'tidb', 'oceanbase', 'greatdb']:
            explain_sql = f"EXPLAIN FORMAT=JSON {sql}"
        elif db_type == 'oracle':
            explain_sql = f"EXPLAIN PLAN FOR {sql}"
        elif db_type == 'shentong':  # 神通数据库兼容Oracle
            explain_sql = f"EXPLAIN PLAN FOR {sql}"
        elif db_type == 'dm':  # 达梦数据库
            explain_sql = f"EXPLAIN {sql}"
        else:
            return jsonify({"status": "error", "message": f"不支持的数据库类型：{db_type}"})
            
        # 执行EXPLAIN语句
        with get_connection_func(db_config) as conn:
            cursor = conn.cursor()
            cursor.execute(explain_sql)
            rows = cursor.fetchall()
            
            # 获取列名
            columns = [desc[0] for desc in cursor.description] if cursor.description else []
            
            # 根据数据库类型处理结果
            plan_data = []
            if db_type == 'oracle' or db_type == 'shentong':
                # Oracle的执行计划需要查询PLAN_TABLE获取详细信息
                # 首先需要确保执行了EXPLAIN PLAN语句，它已经在上面的cursor.execute(explain_sql)中执行
                try:
                    # 查询PLAN_TABLE获取执行计划
                    cursor.execute("SELECT ID, OPERATION, OPTIONS, OBJECT_NAME, OPTIMIZER, COST, CARDINALITY FROM PLAN_TABLE WHERE STATEMENT_ID = (SELECT MAX(STATEMENT_ID) FROM PLAN_TABLE) ORDER BY ID")
                    oracle_plan_rows = cursor.fetchall()
                    columns = [desc[0] for desc in cursor.description] if cursor.description else ["ID", "OPERATION", "OPTIONS", "OBJECT_NAME", "OPTIMIZER", "COST", "CARDINALITY"]
                    plan_data = [dict(zip(columns, row)) for row in oracle_plan_rows]
                    # 清理PLAN_TABLE中的临时数据
                    cursor.execute("DELETE FROM PLAN_TABLE WHERE STATEMENT_ID = (SELECT MAX(STATEMENT_ID) FROM PLAN_TABLE)")
                except Exception as e:
                    # 如果PLAN_TABLE查询失败，返回原始EXPLAIN PLAN的结果
                    plan_data = [dict(zip(columns, row)) for row in rows]
            elif db_type in ['postgresql', 'kingbase', 'highgo', 'gauss', 'uxdb', 'vastbase', 'yashandb', 'gbase', 'vanward']:
                # PostgreSQL的EXPLAIN ANALYZE结果通常包含执行顺序信息，确保以适当格式返回
                # 对于PostgreSQL，结果通常是单列或多列，取决于EXPLAIN选项
                if len(rows) > 0 and len(columns) == 1 and 'QUERY PLAN' in [col.upper() for col in columns]:
                    # 如果是标准的单列查询计划，按行分割并添加统一的执行顺序编号
                    plan_lines = []
                    step_counter = 1  # 统一的步骤计数器
                    for idx, row in enumerate(rows):
                        line_content = str(row[0])
                        # 将多行查询计划分割成单独的行
                        lines = line_content.split('\n')
                        for line in lines:
                            if line.strip():  # 忽略空行
                                plan_lines.append({'Step': step_counter, 'Operation': line.strip()})
                                step_counter += 1
                    if plan_lines:
                        plan_data = plan_lines
                        columns = ['Step', 'Operation']
                    else:
                        # 如果分割后没有内容，使用原来的处理方式
                        plan_data = [dict(zip(['Step', 'Description'], [idx+1, row[0]])) for idx, row in enumerate(rows)]
                        columns = ['Step', 'Description']
                else:
                    # 如果已有多个列，包含执行顺序信息，则添加统一的步骤编号
                    plan_data = [dict(zip(['Step'] + columns, [idx+1] + list(row))) for idx, row in enumerate(rows)]
                    columns = ['Step'] + columns
            elif db_type in ['mysql', 'tidb', 'oceanbase', 'greatdb']:
                # MySQL的EXPLAIN FORMAT=JSON结果可能需要特殊处理
                # 添加统一的步骤编号
                plan_data = []
                for idx, row in enumerate(rows):
                    plan_data.append(dict(zip(['Step'] + columns, [idx+1] + list(row))))
                columns = ['Step'] + columns
                # MySQL JSON格式的执行计划通常在第一行包含完整的JSON格式执行计划
                if plan_data and len(plan_data) > 0:
                    # 确保MySQL的执行计划信息包含顺序相关内容
                    pass  # MySQL的JSON格式本身已经包含了执行计划的层次结构
            else:
                # 其他数据库类型，如达梦
                # 添加统一的步骤编号
                plan_data = []
                for idx, row in enumerate(rows):
                    plan_data.append(dict(zip(['Step'] + columns, [idx+1] + list(row))))
                columns = ['Step'] + columns
                
            return jsonify({
                "status": "success",
                "data": plan_data,
                "columns": columns,
                "db_type": db_type,
                "message": "查询计划分析完成！"
            })
            
    except Exception as e:
        logging.error(f"查询计划分析失败：{str(e)} | SQL：{sql[:100]}...")
        return jsonify({"status": "error", "message": f"查询计划分析失败：{str(e)}"})


def execute_single_statement(sql, page, page_size, db_id):
    """执行单条SQL语句"""
    # 安全校验
    is_safe, msg = check_sql_safety(sql)
    if not is_safe:
        logging.warning(f"SQL安全校验失败：{sql[:100]}... | 原因：{msg}")
        return {"status": "error", "message": msg}
    
    # 获取数据库配置
    db_config = None
    if db_id:
        db_config = get_database_by_id(db_id)
    else:
        db_config = get_default_database()
    
    if not db_config:
        return {"status": "error", "message": "未找到有效的数据库配置"}
    
    db_type = db_config.get('type', 'postgresql').lower()
    
    # 根据数据库类型建立连接并执行SQL
    if db_type == 'postgresql':
        import psycopg2
        from psycopg2 import OperationalError, ProgrammingError
        # 解密密码
        decrypted_password = decrypt_password(db_config.get('password', ''))
        conn = psycopg2.connect(
            host=db_config['host'],
            port=int(db_config['port']),
            user=db_config['user'],
            password=decrypted_password,
            database=db_config['database'],
            connect_timeout=DB_TIMEOUT_CONFIG['connect_timeout']
        )
        # 设置语句超时
        with conn.cursor() as cur:
            cur.execute(f"SET statement_timeout = {DB_TIMEOUT_CONFIG['statement_timeout']}000;")  # 转换为毫秒
    elif db_type == 'mysql':
        import pymysql
        # 解密密码
        decrypted_password = decrypt_password(db_config.get('password', ''))
        conn = pymysql.connect(
            host=db_config['host'],
            port=int(db_config['port']),
            user=db_config['user'],
            password=decrypted_password,
            database=db_config['database'],
            charset='utf8mb4',
            connect_timeout=DB_TIMEOUT_CONFIG['connect_timeout']
        )
        # 设置语句超时
        with conn.cursor() as cur:
            cur.execute(f"SET SESSION MAX_EXECUTION_TIME = {DB_TIMEOUT_CONFIG['statement_timeout']}000;")  # 转换为毫秒
    elif db_type == 'oracle':
        import cx_Oracle
        # 解密密码
        decrypted_password = decrypt_password(db_config.get('password', ''))
        dsn = cx_Oracle.makedsn(db_config['host'], int(db_config['port']), service_name=db_config['database'])
        conn = cx_Oracle.connect(
            user=db_config['user'],
            password=decrypted_password,
            dsn=dsn,
            encoding="UTF-8",
            timeout=DB_TIMEOUT_CONFIG['connect_timeout']
        )
        # 设置语句超时
        with conn.cursor() as cur:
            # Oracle的超时设置在连接级别，通过timeout参数实现
            pass  # Oracle连接级别的timeout参数已经设置
    elif db_type == 'kingbase':  # 人大金仓数据库
        import psycopg2 as kingbase_psycopg2
        # 解密密码
        decrypted_password = decrypt_password(db_config.get('password', ''))
        conn = kingbase_psycopg2.connect(
            host=db_config['host'],
            port=int(db_config['port']),
            user=db_config['user'],
            password=decrypted_password,
            database=db_config['database'],
            connect_timeout=DB_TIMEOUT_CONFIG['connect_timeout']
        )
        # 设置语句超时
        with conn.cursor() as cur:
            cur.execute(f"SET statement_timeout = {DB_TIMEOUT_CONFIG['statement_timeout']}000;")  # 转换为毫秒
    elif db_type in ['tidb', 'oceanbase']:  # TiDB和OceanBase兼容MySQL协议
        import pymysql
        # 解密密码
        decrypted_password = decrypt_password(db_config.get('password', ''))
        conn = pymysql.connect(
            host=db_config['host'],
            port=int(db_config['port']),
            user=db_config['user'],
            password=decrypted_password,
            database=db_config['database'],
            charset='utf8mb4',
            connect_timeout=DB_TIMEOUT_CONFIG['connect_timeout']
        )
        # 设置语句超时
        with conn.cursor() as cur:
            cur.execute(f"SET SESSION MAX_EXECUTION_TIME = {DB_TIMEOUT_CONFIG['statement_timeout']}000;")  # 转换为毫秒
    elif db_type == 'yashandb':  # 崖山数据库使用专用yasdb驱动
        import yasdb
        # 解密密码
        decrypted_password = decrypt_password(db_config.get('password', ''))
        # 构造DSN连接字符串
        dsn = f"{db_config['host']}:{db_config['port']}"
        conn = yasdb.connect(
            dsn=dsn,
            user=db_config['user'],
            password=decrypted_password,
        )
        # 崖山数据库不需要设置语句超时
        pass
    else:
        return {"status": "error", "message": f"不支持的数据库类型：{db_type}"}
    
    try:
        cursor = conn.cursor()
        cursor.execute(sql)
        
        # 处理查询结果
        if cursor.description:
            columns = [desc[0] for desc in cursor.description]
            # 先获取全量结果（内存分页）
            full_results = cursor.fetchall()
            total_count = len(full_results)
            
            # 分页处理
            start = (page - 1) * page_size
            end = start + page_size
            results = full_results[start:end]
            
            # 生成唯一标识，存储查询结果（解决并发问题）
            query_id = str(uuid.uuid4())
            QUERY_RESULTS[query_id] = {
                'columns': columns,
                'results': full_results,
                'create_time': time.time()
            }
            
            data = {
                "status": "success",
                "columns": columns,
                "results": results,
                "count": len(results),
                "total_count": total_count,
                "page": page,
                "page_size": page_size,
                "total_page": (total_count + page_size - 1) // page_size,
                "query_id": query_id  # 返回查询ID用于导出
            }
            logging.info(f"SQL执行成功：{sql[:100]}... | 总记录数：{total_count} | 查询ID：{query_id} | 数据库：{db_id or 'default'}")
        else:
            conn.commit()
            data = {
                "status": "success",
                "message": f"SQL执行成功！影响行数：{cursor.rowcount}"
            }
            logging.info(f"SQL执行成功（非查询）：{sql[:100]}... | 影响行数：{cursor.rowcount} | 数据库：{db_id or 'default'}")
        
        cursor.close()
        return data
    except Exception as e:
        # 详细记录错误信息，包括SQL语句和错误详情
        error_msg = str(e)
        logging.error(f"SQL执行错误 - 时间: {time.strftime('%Y-%m-%d %H:%M:%S')}, DB: {db_id or 'default'}, SQL: {sql}, 错误: {error_msg}", exc_info=True)
        
        # 根据错误类型返回不同的错误信息
        if 'syntax error' in error_msg.lower() or 'parser' in error_msg.lower():
            return {"status": "error", "message": f"SQL语法错误: {str(e)[:200]}..."}
        elif 'permission denied' in error_msg.lower() or 'access denied' in error_msg.lower():
            return {"status": "error", "message": "数据库权限不足，无法执行该操作！"}
        elif 'timeout' in error_msg.lower():
            return {"status": "error", "message": "SQL执行超时，请检查查询语句或联系管理员！"}
        else:
            return {"status": "error", "message": f"SQL执行失败: {str(e)[:200]}..."}
    finally:
        conn.close()

@app.route('/export_excel')
def export_excel():
    """导出Excel（支持自定义表头颜色和是否导出列名）"""
    try:
        # 获取查询ID
        query_id = request.args.get('query_id')
        if not query_id or query_id not in QUERY_RESULTS:
            return jsonify({"status": "error", "message": "查询结果不存在或已过期！请重新执行查询。"})
        
        # 获取Excel配置参数
        header_color = request.args.get('header_color', '4472C4')
        include_header = request.args.get('include_header', 'true').lower() == 'true'
        
        # 获取自定义文件名（如果提供）
        custom_filename = request.args.get('filename', '')
        if custom_filename:
            filename = custom_filename if custom_filename.endswith('.xlsx') else f"{custom_filename}.xlsx"
        else:
            filename = f"SQL查询结果_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        
        # 获取全量结果
        last_result = QUERY_RESULTS[query_id]
        columns = last_result['columns']
        results = last_result['results']
        
        # 创建Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"查询结果_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        # 写入表头（如果需要）
        if include_header and columns:
            ws.append(columns)
        
        # 写入数据
        for row in results:
            ws.append(row)
        
        # 设置样式（只有存在表头时才设置表头样式）
        if include_header and columns:
            set_excel_style(ws, header_color)
        
        # 保存到BytesIO
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        logging.info(f"Excel导出成功：查询ID={query_id} | 记录数：{len(results)} | 表头颜色：{header_color} | 包含表头：{include_header} | 文件名：{filename}")
        
        # 下载文件
        return send_file(
            excel_bytes,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        logging.error(f"Excel导出失败：{str(e)}")
        return jsonify({"status": "error", "message": f"导出Excel失败：{str(e)}"})

@app.route('/export_csv')
def export_csv():
    """导出CSV（支持自定义分隔符和是否导出列名）"""
    try:
        # 获取查询ID
        query_id = request.args.get('query_id')
        if not query_id or query_id not in QUERY_RESULTS:
            return jsonify({"status": "error", "message": "查询结果不存在或已过期！请重新执行查询。"})
        
        # 获取CSV配置参数
        separator_type = request.args.get('separator', 'comma')
        include_header = request.args.get('include_header', 'true').lower() == 'true'
        
        # 获取分隔符
        separator = SUPPORTED_CSV_SEPARATORS.get(separator_type, ',')
        
        # 获取自定义文件名（如果提供）
        custom_filename = request.args.get('filename', '')
        if custom_filename:
            filename = custom_filename if custom_filename.endswith('.csv') else f"{custom_filename}.csv"
        else:
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"SQL查询结果_{query_id}_{timestamp}.csv"
        
        # 获取全量结果
        last_result = QUERY_RESULTS[query_id]
        columns = last_result['columns']
        results = last_result['results']
        
        # 生成CSV内容
        csv_content = generate_csv_content(columns, results, separator, include_header)
        
        logging.info(f"CSV导出成功：{filename} | 记录数：{len(results)} | 分隔符：{repr(separator)} | 包含表头：{include_header}")
        
        # 返回CSV文件
        return send_file(
            csv_content,
            as_attachment=True,
            download_name=filename,
            mimetype='text/csv; charset=utf-8-sig'
        )
    
    except Exception as e:
        logging.error(f"CSV导出失败：{str(e)}")
        return jsonify({"status": "error", "message": f"导出CSV失败：{str(e)}"})


@app.route('/export_html')
def export_html():
    """导出HTML（类似Oracle AWR报告样式）"""
    try:
        # 获取查询ID
        query_id = request.args.get('query_id')
        if not query_id or query_id not in QUERY_RESULTS:
            return jsonify({"status": "error", "message": "查询结果不存在或已过期！请重新执行查询。"})
        
        # 获取自定义文件名（如果提供）
        custom_filename = request.args.get('filename', '')
        if custom_filename:
            filename = custom_filename if custom_filename.endswith('.html') else f"{custom_filename}.html"
        else:
            filename = f"SQL查询结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        
        # 获取全量结果
        last_result = QUERY_RESULTS[query_id]
        columns = last_result['columns']
        results = last_result['results']
        
        # 生成HTML内容
        html_content = generate_awr_style_html(columns, results)
        
        # 创建临时文件并返回
        from io import StringIO
        import tempfile
        import os
        
        # 创建临时文件
        temp_file = StringIO(html_content)
        
        from flask import send_file
        import io
        
        # 将HTML内容转换为字节流
        html_bytes = html_content.encode('utf-8')
        byte_io = io.BytesIO(html_bytes)
        
        return send_file(
            byte_io,
            as_attachment=True,
            download_name=filename,
            mimetype='text/html; charset=utf-8'
        )
    
    except Exception as e:
        logging.error(f"HTML导出失败：{str(e)}")
        return jsonify({"status": "error", "message": f"导出HTML失败：{str(e)}"})

def generate_awr_style_html(columns, results):
    """生成类似Oracle AWR报告样式的HTML"""
    # HTML头部，包含CSS样式
    html_head = '''<!DOCTYPE HTML PUBLIC "-//W3C//DTD HTML 4.01 Transitional//EN" "http://www.w3.org/TR/html4/loose.dtd">
<html>
<head>
    <meta http-equiv="Content-Type" content="text/html; charset=UTF-8">
    <title>SQL查询结果报告 - Oracle AWR风格</title>
    <style type="text/css">
        body {font:9pt Arial,Helvetica,sans-serif; color:black; background:White;}
        p {font:9pt Arial,Helvetica,sans-serif; color:black; background:White;}
        table,tr,td {font:9pt Arial,Helvetica,sans-serif; color:Black; background:#FFFFCC; padding:0px 0px 0px 0px; margin:0px 0px 0px 0px;}
        th {font:bold 9pt Arial,Helvetica,sans-serif; color:White; background:#0066CC; padding:0px 0px 0px 0px;}
        h1 {font:bold 12pt Arial,Helvetica,Geneva,sans-serif; color:#336699; background-color:White; border-bottom:1px solid #cccc99; margin-top:0pt; margin-bottom:0pt; padding:0px 0px 0px 0px;}
        a {font:10pt Arial,Helvetica,sans-serif; color:#0066CC; margin-top:0pt; margin-bottom:0pt; vertical-align:top;text-decoration: none;}
        a.link {font:10pt Arial,Helvetica,sans-serif; color:#0066CC; margin-top:0pt; margin-bottom:0pt; vertical-align:top;text-decoration: none;}
        .awr-report { background:white; padding:0px; margin:0px;}
        .report-header { font:bold 14pt Arial,Helvetica,Geneva,sans-serif; color:#336699; background-color:White; border-bottom:1px solid #cccc99; margin-top:0pt; margin-bottom:0pt; padding:10px 0px 10px 0px; text-align:center;}
        .section-title { font:bold 10pt Arial,Helvetica,Geneva,sans-serif; color:white; background:#0066CC; margin:15px 0px 5px 0px; padding:5px;}
        .summary-info { font:9pt Arial,Helvetica,sans-serif; margin:10px 0px 10px 0px;}
        .summary-item { margin:5px 0px 5px 0px;}
        .report-footer { font:8pt Arial,Helvetica,sans-serif; color:#666666; text-align:center; margin-top:20px; padding-top:5px; border-top:1px solid #cccc99;}
    </style>
</head>
<body BGCOLOR="#C0C0C0">
    <div class="awr-report">
        <h1 class="report-header">SQL查询结果报告</h1>
        <p class="summary-info">
            <span class="summary-item"><b>报告生成时间:</b> ''' + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + '''</span><br>
            <span class="summary-item"><b>总记录数:</b> ''' + str(len(results)) + '''</span><br>
            <span class="summary-item"><b>总列数:</b> ''' + str(len(columns)) + '''</span><br>
        </p>
        <div class="section-title">查询结果</div>
        <table WIDTH="100%" CELLPADDING="2" CELLSPACING="0" BORDER="1" BORDERCOLOR="#0066CC" BGCOLOR="#FFFFCC">
            <thead>
                <tr>'''
    
    # 表头
    html_body_start = ''
    for col in columns:
        html_body_start += f'<th>{col}</th>'
    
    # 表头结束和表体开始
    html_body_middle = '''</tr>
            </thead>
            <tbody>'''
    
    # 表格内容
    html_body_end = ''
    for i, row in enumerate(results):
        html_body_end += '<tr>' 
        for cell in row:
            cell_value = str(cell) if cell is not None else '&nbsp;'
            html_body_end += f'<td>{cell_value}</td>'
        html_body_end += '</tr>'
    
    # HTML底部
    html_foot = '''</tbody>
        </table>
        <div class="report-footer">
            Generated by 数据报表工具 | ''' + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + '''
        </div>
    </div>
</body>
</html>'''

    
    # 合并所有部分
    return html_head + html_body_start + html_body_middle + html_body_end + html_foot

@app.route('/set_default_db', methods=['POST'])
def set_default_db():
    """设置默认数据库"""
    try:
        data = request.json
        db_id = data.get('db_id')
        
        if not db_id:
            return jsonify({"status": "error", "message": "数据库ID不能为空！"})
        
        success = set_default_database(db_id)
        if success:
            return jsonify({"status": "success", "message": "默认数据库设置成功！"})
        else:
            return jsonify({"status": "error", "message": "设置默认数据库失败！"})
    except Exception as e:
        logging.error(f"设置默认数据库失败：{str(e)}")
        return jsonify({"status": "error", "message": f"设置失败：{str(e)}"})


@app.route('/get_saved_db_config')
def get_saved_db_config():
    """获取保存的数据库配置文件"""
    try:
        config = load_multi_db_config()
        databases = config.get('databases', [])
        
        # 返回数据库配置列表（不包含密码）
        result = []
        for db in databases:
            db_copy = db.copy()
            # 移除密码字段以确保安全
            if 'password' in db_copy:
                del db_copy['password']
            result.append(db_copy)
        
        return jsonify(result)
    except Exception as e:
        logging.error(f"获取数据库配置失败：{str(e)}")
        return jsonify([])  # 返回空数组而不是错误

        
        # 返回数据库配置列表（包含加密密码）
        result = []
        for db in databases:
            db_copy = db.copy()
            # 保持密码字段（已加密存储）
            result.append(db_copy)
        
        return jsonify(result)
    except Exception as e:
        logging.error(f"获取数据库配置失败：{str(e)}")
        return jsonify([])  # 返回空数组而不是错误


@app.route('/databases', methods=['GET', 'POST', 'PUT', 'DELETE'])
def manage_databases():
    """数据库配置管理接口：GET获取列表，POST新增，PUT更新，DELETE删除"""
    try:
        if request.method == 'GET':
            # 获取数据库配置列表
            config = load_multi_db_config()
            databases = config.get('databases', [])
            # 不返回密码字段
            for db in databases:
                if 'password' in db:
                    del db['password']
            return jsonify({"status": "success", "data": databases})
        
        elif request.method == 'POST':
            # 新增数据库配置
            data = request.json or {}
            name = (data.get('name') or '').strip()
            db_type = (data.get('type') or 'postgresql').lower()
            host = (data.get('host') or '').strip()
            port = str(data.get('port', '')).strip()
            user = (data.get('user') or '').strip()
            password = data.get('password', '')
            database = (data.get('database') or '').strip()
            
            if not name or not host or not port or not user or not database:
                return jsonify({"status": "error", "message": "数据库名称、主机、端口、用户名和数据库名为必填项！"})
            
            if db_type not in SUPPORTED_DATABASES:
                return jsonify({"status": "error", "message": f"不支持的数据库类型：{db_type}"})
            
            # 读取现有配置
            try:
                with open(DB_CONFIG_FILE, 'r', encoding='utf-8') as f:
                    existing_config = json.load(f)
            except FileNotFoundError:
                existing_config = {"databases": []}
            except json.JSONDecodeError:
                existing_config = {"databases": []}
            
            # 检查是否已存在同名数据库
            for db in existing_config.get('databases', []):
                if db['name'] == name:
                    return jsonify({"status": "error", "message": "数据库名称已存在！"})
            
            # 创建新数据库配置
            new_db = {
                "id": str(uuid.uuid4()),
                "name": name,
                "type": db_type,
                "host": host,
                "port": port,
                "user": user,
                "password": encrypt_password(password),
                "database": database,
                "is_default": False  # 新增时不设为默认
            }
            
            existing_config.setdefault('databases', []).append(new_db)
            
            with open(DB_CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(existing_config, f, ensure_ascii=False, indent=2)
            
            logging.info(f"新增数据库配置：{name}")
            return jsonify({"status": "success", "message": "数据库配置新增成功！"})
        
        elif request.method == 'PUT':
            with DB_CONFIG_LOCK:  # 添加锁
                # 更新数据库配置
                data = request.json or {}
                db_id = data.get('id')
                name = (data.get('name') or '').strip()
                db_type = (data.get('type') or 'postgresql').lower()
                host = (data.get('host') or '').strip()
                port = (data.get('port') or '').strip()
                user = (data.get('user') or '').strip()
                password = data.get('password', '')
                database = (data.get('database') or '').strip()
                
                if not db_id or not name or not host or not port or not user or not database:
                    return jsonify({"status": "error", "message": "ID、名称、主机、端口、用户名和数据库名为必填项！"})
                
                if db_type not in SUPPORTED_DATABASES:
                    return jsonify({"status": "error", "message": f"不支持的数据库类型：{db_type}"})
                
                # 读取现有配置
                try:
                    with open(DB_CONFIG_FILE, 'r', encoding='utf-8') as f:
                        existing_config = json.load(f)
                except FileNotFoundError:
                    return jsonify({"status": "error", "message": "数据库配置文件不存在！"})
                except json.JSONDecodeError:
                    return jsonify({"status": "error", "message": "数据库配置文件格式错误！"})
                
                # 查找并更新数据库配置
                databases = existing_config.get('databases', [])
                db_found = False
                for i, db in enumerate(databases):
                    if db['id'] == db_id:
                        # 保留是否为默认数据库的状态
                        is_default = db.get('is_default', False)
                        
                        # 获取原始密码
                        original_password = db.get('password', '')
                        
                        # 处理密码字段：只有当明确提供新密码时才更新
                        # 如果没有提供密码字段，或密码字段为空，或与原密码相同，则保留原密码
                        if 'password' in data and data.get('password'):
                            # 检查是否是新密码（不等于原始密码且不是加密格式）
                            provided_password = data.get('password')
                            if provided_password != original_password and not provided_password.startswith('enc:'):
                                # 提供了新密码，进行加密
                                updated_password = encrypt_password(provided_password)
                            else:
                                # 提供的密码与原密码相同或已是加密格式，保留原密码
                                updated_password = original_password
                        else:
                            # 没有提供密码或密码为空，保留原始密码
                            updated_password = original_password
                        
                        updated_db = {
                            "id": db_id,
                            "name": name,
                            "type": db_type,
                            "host": host,
                            "port": port,
                            "user": user,
                            "password": updated_password,
                            "database": database,
                            "is_default": is_default
                        }
                        databases[i] = updated_db
                        db_found = True
                        break
                
                if not db_found:
                    return jsonify({"status": "error", "message": "未找到对应的数据库配置！"})
            
            with open(DB_CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(existing_config, f, ensure_ascii=False, indent=2)
            
            logging.info(f"更新数据库配置：{name}")
            return jsonify({"status": "success", "message": "数据库配置更新成功！"})
        
        elif request.method == 'DELETE':
            with DB_CONFIG_LOCK:  # 添加锁
                # 删除数据库配置
                db_id = request.args.get('id')
                if not db_id:
                    return jsonify({"status": "error", "message": "缺少ID参数！"})
                
                # 读取现有配置
                try:
                    with open(DB_CONFIG_FILE, 'r', encoding='utf-8') as f:
                        existing_config = json.load(f)
                except FileNotFoundError:
                    return jsonify({"status": "error", "message": "数据库配置文件不存在！"})
                except json.JSONDecodeError:
                    return jsonify({"status": "error", "message": "数据库配置文件格式错误！"})
                
                # 查找并删除数据库配置
                databases = existing_config.get('databases', [])
                new_databases = [db for db in databases if db['id'] != db_id]
                
                if len(new_databases) == len(databases):
                    return jsonify({"status": "error", "message": "未找到对应的数据库配置！"})
                
                existing_config['databases'] = new_databases
                
                with open(DB_CONFIG_FILE, 'w', encoding='utf-8') as f:
                    json.dump(existing_config, f, ensure_ascii=False, indent=2)
                
                logging.info(f"删除数据库配置：{db_id}")
                return jsonify({"status": "success", "message": "数据库配置删除成功！"})
            
    except Exception as e:
        logging.error(f"数据库配置管理接口执行失败：{e}")
        return jsonify({"status": "error", "message": f"操作失败：{str(e)}"})


# ===================== 启动 =====================



def load_raw_db_config():
    """加载原始数据库配置文件（不解密密码）"""
    try:
        with open(DB_CONFIG_FILE, 'r', encoding='utf-8') as f:
            config = json.load(f)
        return config
    except Exception as e:
        logging.error(f"加载原始数据库配置失败：{str(e)}")
        return {"databases": []}


@app.route('/get_saved_db_config_with_password')
def get_saved_db_config_with_password():
    """获取保存的数据库配置文件（包含加密密码）"""
    try:
        config = load_raw_db_config()
        databases = config.get('databases', [])
        
        # 返回数据库配置列表（包含加密密码）
        result = []
        for db in databases:
            db_copy = db.copy()
            # 保持密码字段（已加密存储）
            result.append(db_copy)
        
        return jsonify(result)
    except Exception as e:
        logging.error(f"获取数据库配置失败：{str(e)}")
        return jsonify([])  # 返回空数组而不是错误


@app.route('/get_db_config_with_password/<db_id>')
def get_db_config_with_password(db_id):
    """根据ID获取单个数据库配置（包含加密密码）"""
    try:
        config = load_raw_db_config()
        databases = config.get('databases', [])
        
        # 查找指定ID的数据库配置
        for db in databases:
            if db.get('id') == db_id:
                # 返回该数据库配置（包含加密密码）
                logging.info(f"DEBUG: Returning database config for ID {db_id} with password field present: {'password' in db}")
                return jsonify({"status": "success", "data": db})
        
        return jsonify({"status": "error", "message": "未找到指定的数据库配置"})
    except Exception as e:
        logging.error(f"获取数据库配置失败：{str(e)}")
        return jsonify({"status": "error", "message": f"获取失败：{str(e)}"})


@app.route('/check_app_password', methods=['POST'])
def check_app_password():
    """验证应用访问密码"""
    try:
        data = request.json
        input_password = data.get('password', '')
        
        # 加载已保存的密码
        saved_password = load_app_password()
        
        # 验证密码
        if saved_password == "" or input_password == saved_password:
            # 如果没有设置密码，或者输入密码正确，则验证通过
            # 返回会话令牌用于后续API调用
            session_token = hashlib.sha256(saved_password.encode()).hexdigest() if saved_password else ""
            return jsonify({
                "status": "success", 
                "message": "验证通过", 
                "has_password": saved_password != "",
                "session_token": session_token
            })
        else:
            # 密码错误
            return jsonify({"status": "error", "message": "密码错误"})
    except Exception as e:
        logging.error(f"验证应用访问密码失败：{str(e)}")
        return jsonify({"status": "error", "message": f"验证失败：{str(e)}"})


@app.route('/change_app_password', methods=['POST'])
def change_app_password():
    """更改应用访问密码（需要验证旧密码）"""
    try:
        data = request.json
        old_password = data.get('old_password', '')
        new_password = data.get('new_password', '')
        confirm_new_password = data.get('confirm_new_password', '')
        
        # 验证新密码和确认密码是否一致
        if new_password != confirm_new_password:
            return jsonify({"status": "error", "message": "新密码和确认密码不一致"})
        
        # 验证新密码长度
        if len(new_password) < 1:
            return jsonify({"status": "error", "message": "新密码不能为空"})
        
        # 检查是否启用了密码强度验证
        password_strength_required = APP_CONFIG.get('app_password_strength_required', False)
        if password_strength_required:
            # 验证密码强度
            if not is_strong_password(new_password):
                return jsonify({"status": "error", "message": "密码不符合强度要求：至少8位，包含大写字母、小写字母、数字和特殊字符"})
        
        # 加载当前保存的密码
        current_password = load_app_password()
        
        # 如果当前没有设置密码，则直接设置新密码
        if current_password == "":
            success = save_app_password(new_password)
            if success:
                return jsonify({"status": "success", "message": "密码设置成功"})
            else:
                return jsonify({"status": "error", "message": "密码设置失败"})
        else:
            # 如果当前已有密码，则需要验证旧密码
            if old_password != current_password:
                return jsonify({"status": "error", "message": "旧密码错误"})
            
            # 保存新密码
            success = save_app_password(new_password)
            
            if success:
                return jsonify({"status": "success", "message": "密码修改成功"})
            else:
                return jsonify({"status": "error", "message": "密码修改失败"})
    except Exception as e:
        logging.error(f"更改应用访问密码失败：{str(e)}")
        return jsonify({"status": "error", "message": f"修改失败：{str(e)}"})


@app.route('/set_app_password', methods=['POST'])
def set_app_password():
    """设置应用访问密码"""
    try:
        data = request.json
        new_password = data.get('password', '')
        confirm_password = data.get('confirm_password', '')
        
        # 验证密码和确认密码是否一致
        if new_password != confirm_password:
            return jsonify({"status": "error", "message": "密码和确认密码不一致"})
        
        # 验证密码长度
        if len(new_password) < 1:
            return jsonify({"status": "error", "message": "密码不能为空"})
        
        # 检查是否启用了密码强度验证
        password_strength_required = APP_CONFIG.get('app_password_strength_required', False)
        if password_strength_required:
            # 验证密码强度
            if not is_strong_password(new_password):
                return jsonify({"status": "error", "message": "密码不符合强度要求：至少8位，包含大写字母、小写字母、数字和特殊字符"})
        
        # 保存密码（加密存储）
        success = save_app_password(new_password)
        
        if success:
            return jsonify({"status": "success", "message": "密码设置成功"})
        else:
            return jsonify({"status": "error", "message": "密码设置失败"})
    except Exception as e:
        logging.error(f"设置应用访问密码失败：{str(e)}")
        return jsonify({"status": "error", "message": f"设置失败：{str(e)}"})


@app.route('/has_app_password')
def has_app_password():
    """检查是否设置了应用访问密码"""
    try:
        saved_password = load_app_password()
        has_password = saved_password != ""
        return jsonify({"has_password": has_password})
    except Exception as e:
        logging.error(f"检查应用访问密码状态失败：{str(e)}")
        return jsonify({"has_password": False})


if __name__ == '__main__':
    import sys
    port = 5000  # 默认端口
    for arg in sys.argv:
        if arg.startswith('--port='):
            try:
                port = int(arg.split('=')[1])
                break
            except ValueError:
                print(f"无效的端口号: {arg}")
                sys.exit(1)

    @app.route('/get_app_config')
    def get_app_config():
        """获取应用配置"""
        return jsonify({
            "status": "success",
            "config": APP_CONFIG
        })


    @app.route('/save_app_config', methods=['POST'])
    @require_auth
    def save_app_config():
        """保存应用配置"""
        # 在函数开始时声明全局变量
        global APP_CONFIG, DB_TIMEOUT_CONFIG
        
        try:
            data = request.get_json()
            
            # 验证输入数据
            if not data:
                return jsonify({"status": "error", "message": "缺少配置数据"}), 400
            
            timeout_minutes = data.get('app_auto_lock_timeout_minutes')
            reminder_minutes = data.get('app_auto_lock_reminder_minutes')
            app_title = data.get('app_title', '朝阳数据')  # 获取应用标题，如果未提供则使用默认值
            statement_timeout = data.get('db_statement_timeout', 30)  # 数据库语句超时时间
            connect_timeout = data.get('db_connect_timeout', 10)      # 数据库连接超时时间
            app_password = data.get('app_password', '')              # 应用访问密码
            max_connections = data.get('app_max_connections', 10)    # 最大连接数
            min_connections = data.get('app_min_connections', 1)     # 最小连接数
            connection_pool_timeout = data.get('app_connection_pool_timeout', 30)  # 连接池超时时间
            result_cache_time = data.get('app_result_cache_time', 3600)           # 结果缓存时间
            max_result_size = data.get('app_max_result_size', 10000)              # 最大结果集大小
            login_failures_limit = data.get('app_login_failures_limit', 5)        # 登录失败次数限制
            account_lockout_minutes = data.get('app_account_lockout_minutes', 30) # 账户锁定时间
            password_strength_required = data.get('app_password_strength_required', False)  # 密码强度要求
            log_level = data.get('app_log_level', 'INFO')                       # 日志级别
            log_retention_days = data.get('app_log_retention_days', 30)         # 日志保留天数
            audit_logging_enabled = data.get('app_audit_logging_enabled', True)  # 审计日志开关
            theme_color = data.get('app_theme_color', 'default')               # 主题色
            language = data.get('app_language', 'zh-CN')                      # 界面语言
            page_size = data.get('app_page_size', 50)                        # 页面大小
            auto_save_interval = data.get('app_auto_save_interval', 300)      # 自动保存间隔
            concurrent_queries = data.get('app_concurrent_queries', 5)       # 并发查询数
            query_queue_size = data.get('app_query_queue_size', 10)         # 查询队列大小
            memory_limit_mb = data.get('app_memory_limit_mb', 512)         # 内存限制
            batch_insert_size = data.get('app_batch_insert_size', 1000)   # 批量插入大小
            transaction_timeout = data.get('app_transaction_timeout', 120)  # 事务超时时间
            connection_retry_count = data.get('app_connection_retry_count', 3)  # 连接重试次数
            
            if timeout_minutes is None or reminder_minutes is None:
                return jsonify({"status": "error", "message": "缺少必要的配置参数"}), 400
            
            # 验证配置值的合理性
            if not isinstance(timeout_minutes, (int, float)) or not isinstance(reminder_minutes, (int, float)):
                return jsonify({"status": "error", "message": "配置值必须为数字"}), 400
            
            if timeout_minutes <= 0 or timeout_minutes > 1440:
                return jsonify({"status": "error", "message": "自动锁定时间应在1-1440分钟之间"}), 400
            
            if reminder_minutes <= 0 or reminder_minutes >= timeout_minutes:
                return jsonify({"status": "error", "message": "提醒时间应大于0且小于自动锁定时间"}), 400
            
            # 验证应用标题
            if not isinstance(app_title, str) or len(app_title.strip()) == 0:
                return jsonify({"status": "error", "message": "应用标题不能为空"}), 400
            
            # 限制标题长度
            if len(app_title) > 50:
                return jsonify({"status": "error", "message": "应用标题不能超过50个字符"}), 400
            
            # 验证数据库超时配置
            if not isinstance(statement_timeout, (int, float)) or statement_timeout <= 0:
                return jsonify({"status": "error", "message": "语句超时时间必须为正数"}), 400
                
            if not isinstance(connect_timeout, (int, float)) or connect_timeout <= 0:
                return jsonify({"status": "error", "message": "连接超时时间必须为正数"}), 400
            
            # 创建新的配置字典
            new_config = {
                "app_auto_lock_timeout_minutes": int(timeout_minutes),
                "app_auto_lock_reminder_minutes": int(reminder_minutes),
                "app_title": app_title.strip(),  # 保存提供的标题
                "db_statement_timeout": int(statement_timeout),
                "db_connect_timeout": int(connect_timeout),
                "app_max_connections": int(max_connections),
                "app_min_connections": int(min_connections),
                "app_connection_pool_timeout": int(connection_pool_timeout),
                "app_result_cache_time": int(result_cache_time),
                "app_max_result_size": int(max_result_size),
                "app_login_failures_limit": int(login_failures_limit),
                "app_account_lockout_minutes": int(account_lockout_minutes),
                "app_password_strength_required": bool(password_strength_required),
                "app_log_level": log_level,
                "app_log_retention_days": int(log_retention_days),
                "app_audit_logging_enabled": bool(audit_logging_enabled),
                "app_theme_color": theme_color,
                "app_language": language,
                "app_page_size": int(page_size),
                "app_auto_save_interval": int(auto_save_interval),
                "app_concurrent_queries": int(concurrent_queries),
                "app_query_queue_size": int(query_queue_size),
                "app_memory_limit_mb": int(memory_limit_mb),
                "app_batch_insert_size": int(batch_insert_size),
                "app_transaction_timeout": int(transaction_timeout),
                "app_connection_retry_count": int(connection_retry_count)
            }
            
            # 如果提供了密码，则加密并添加到配置中
            if app_password:  # 只有当密码不为空时才更新
                # 检查是否启用了密码强度验证
                if password_strength_required:
                    # 验证密码强度
                    if not is_strong_password(app_password):
                        return jsonify({"status": "error", "message": "密码不符合强度要求：至少8位，包含大写字母、小写字母、数字和特殊字符"}), 400
                encrypted_password = encrypt_app_password(app_password)
                new_config["app_password"] = encrypted_password
            else:
                # 如果没有提供密码，保留原有密码或使用空密码
                # 先从原配置文件读取当前密码，而不是直接使用内存中的APP_CONFIG
                try:
                    with open(APP_CONFIG_FILE, 'r', encoding='utf-8') as f:
                        current_config = json.load(f)
                    existing_password = current_config.get("app_password", "")
                except:
                    existing_password = ""
                new_config["app_password"] = existing_password
            
            # 保存到配置文件
            with open(APP_CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(new_config, f, ensure_ascii=False, indent=4)
            
            # 更新内存中的配置
            APP_CONFIG = load_app_config()
            
            # 更新数据库超时配置
            DB_TIMEOUT_CONFIG = get_db_timeout_config()
            
            logging.info(f"应用配置已更新: {new_config}")
            
            return jsonify({
                "status": "success",
                "message": "配置保存成功"
            })
            
        except Exception as e:
            logging.error(f"保存应用配置失败: {str(e)}")
            return jsonify({
                "status": "error", 
                "message": f"保存配置失败: {str(e)}"
            }), 500

    logging.info(f"SQL查询工具已启动（支持Excel/CSV/HTML导出），端口：{port}）")
    app.run(host='0.0.0.0', port=port, debug=False)